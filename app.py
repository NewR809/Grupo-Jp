#IMPORTACIONES
import sys, os
sys.path.append(os.path.dirname(os.path.dirname(__file__)))

from db import guardar_en_mysql
import sys, os
sys.path.append(os.path.dirname(os.path.dirname(__file__)))
from dotenv import load_dotenv
load_dotenv()
from db import guardar_en_mysql
import os
import csv
import pandas as pd
from datetime import datetime
from openpyxl.styles import Alignment, Font, PatternFill
import time
import sys
import ttkbootstrap as ttkbcliner
from ttkbootstrap.constants import *
from tkinter import messagebox, simpledialog, Menu, Toplevel, StringVar, ttk, Entry, Text
from PIL import Image, ImageTk
import tkinter as tk
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import sqlite3
import os
import sqlite3
import mysql.connector
from db import guardar_en_mysql
from conexion import crear_conexion
from base_datos import guardar_en_mysql
from archivos import guardar_csv, leer_csv
#3

# from desktop.gastos_handler import procesar_gasto  # Eliminado: módulo no encontrado
#$60
import requests
from db import guardar_en_mysql
from archivos import guardar_csv, leer_csv
from conexion import crear_conexion
from preview_module import VistaPrevia
from vista_previa import VistaPreviaEditable
from ingresos_handler import procesar_ingreso



# --- Configuración de conexión MySQL ---
DB_CONFIG = {
    "host": "gondola.proxy.rlwy.net",
    "user": "root",
    "port": 18615,
    "password": "DKdNBPtQrzWVwArUWDqIFKEzbSnQIvlG",
    "database": "railway"
}


# --- Credenciales ---
usuarios = {
    "Henko01": "Hen4514",
    "Tpack": "Tpack4514",
    "Jpsolar": "Jpsolar4514",
    "Quidopedia": "Qpd4514"
}
administrador = {
    "Admin": "Jpg4514"
}

nombres_usuarios = {
    "Henko01": "Usuario",
    "Tpack": "Usuario",
    "JpSolar": "Usuario",
    "Quidopedia": "Usuario",
    "Admin": "Gerente"
}

CLAVES_EMPRESAS = {
    "Henko01": "henko123",
    "Tpack": "tpack123",
    "JpSolar": "jpsolar123",
    "Quidopedia": "quidopedia123"
}


# --- Archivos ---
GASTOS_FILE = "gastos.csv"
INGRESOS_FILE = "ingresos.csv"
LOG_SESIONES = "log_sesiones.csv"
TIEMPO_MAX_INACTIVIDAD = 300  # segundos

# --- Registrar sesión ---
def registrar_log_sesion(usuario, tipo):
    fecha_hora = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    guardar_csv(LOG_SESIONES, [fecha_hora, usuario, tipo])


# --- Utilidades de archivo ---
def guardar_csv(archivo, datos):
    with open(archivo, mode='a', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(datos)

def leer_csv(archivo):
    if not os.path.exists(archivo):
        return []
    with open(archivo, mode='r', encoding='utf-8') as f:
        return list(csv.reader(f))

def ajustar_margenes_excel(writer, sheet_name, es_admin=False):
    ws = writer.sheets[sheet_name]
    color = "FFCCCC" if es_admin else "CCE5FF"
    for i in range(1, 4):
        for cell in ws[i]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.column_dimensions[column].width = max_length + 2

def exportar_a_excel_consolidado(archivo_gastos, archivo_ingresos, archivo_excel, usuario_actual=None):
    # Función modificada para aceptar DataFrames en lugar de rutas de archivo
    def leer_o_usar_df(data):
        if isinstance(data, str) and os.path.exists(data):
            return pd.read_csv(data, header=None)
        elif isinstance(data, pd.DataFrame):
            return data
        return pd.DataFrame()

    df_gastos = leer_o_usar_df(archivo_gastos)
    df_ingresos = leer_o_usar_df(archivo_ingresos)
    
    if df_gastos.empty and df_ingresos.empty:
        messagebox.showerror("Error", "No hay datos para exportar.")
        return

    fecha_actual = datetime.now().strftime("%d/%m/%Y %H:%M")
    nombre_usuario = nombres_usuarios.get(usuario_actual, usuario_actual) if usuario_actual else "Sistema Financiero"
    es_admin = usuario_actual == "Admin"

    with pd.ExcelWriter(archivo_excel, engine='openpyxl') as writer:
        if not df_gastos.empty:
            encabezado_gastos = pd.DataFrame([[f"Exportado por: {nombre_usuario}", f"Fecha: {fecha_actual}"], ["Reporte: Gastos", ""]])
            encabezado_gastos.to_excel(writer, index=False, header=False, sheet_name="Gastos")
            df_gastos.columns = ["Fecha", "Usuario", "Monto", "Categoría/Fuente", "Descripción"]
            df_gastos.to_excel(writer, index=False, sheet_name="Gastos", startrow=3)
            ajustar_margenes_excel(writer, "Gastos", es_admin)

        if not df_ingresos.empty:
            encabezado_ingresos = pd.DataFrame([[f"Exportado por: {nombre_usuario}", f"Fecha: {fecha_actual}"], ["Reporte: Ingresos", ""]])
            encabezado_ingresos.to_excel(writer, index=False, header=False, sheet_name="Ingresos")
            df_ingresos.columns = ["Fecha", "Usuario", "Monto", "Categoría/Fuente", "Descripción"]
            df_ingresos.to_excel(writer, index=False, sheet_name="Ingresos", startrow=3)
            ajustar_margenes_excel(writer, "Ingresos", es_admin)

    messagebox.showinfo("Exportación exitosa", f"Archivo exportado como: {archivo_excel}")

# --- Vista previa antes de exportar ---




# --- Interfaz con ttkbootstrap ---
class SistemaFinancieroApp:
    def __init__(self, root):
        self.root = root
        #self.usuario = "Admini"
        self.root.title("Sistema Financiero Empresarial")
        self.root.geometry("800x600")
        self.usuario = None
        self.tipo = None
        self.ultima_actividad = time.time()

        self.root.bind_all("<Any-KeyPress>", self.actualizar_actividad)
        self.root.bind_all("<Any-Button>", self.actualizar_actividad)
        self.root.after(10000, self.verificar_inactividad)

        self.crear_menu_barra()
        self.mostrar_login()

    def actualizar_actividad(self, event=None):
        self.ultima_actividad = time.time()

    def verificar_inactividad(self):
        try:
            if time.time() - self.ultima_actividad > TIEMPO_MAX_INACTIVIDAD:
                messagebox.showinfo("Sesión terminada", "Se cerró la sesión por inactividad.")
                self.mostrar_login()
            else:
                # Reprograma la verificación cada 10 segundos
                self.root.after(10000, self.verificar_inactividad)
        except tk.TclError:
            # La ventana principal ya fue destruida, no hacer nada
            pass

    def crear_menu_barra(self):
        self.menubar = Menu(self.root)
        self.root.config(menu=self.menubar)

        self.menu_archivo = Menu(self.menubar, tearoff=0)
        self.menu_archivo.add_command(label="Cerrar sesión", command=self.mostrar_login)
        self.menu_archivo.add_separator()
        self.menu_archivo.add_command(label="Salir", command=self.salir_app)
        self.menubar.add_cascade(label="Archivo", menu=self.menu_archivo)

        # Limpiar menú de admin anterior si existe
        try:
            self.menubar.delete("Administración")
        except tk.TclError:
            pass
        
        if self.tipo == "admin":
            self.menu_admin = Menu(self.menubar, tearoff=0)
            self.menu_admin.add_command(label="Gestión de Empresas", command=self.gestion_empresas)
            self.menubar.add_cascade(label="Administración", menu=self.menu_admin)

    def salir_app(self):
        self.root.after_cancel(self.root)
        self.root.quit()
        sys.exit()

    def crear_barra_tareas(self):
        barra = ttkbcliner.Frame(self.root, bootstyle="dark")
        barra.pack(side="top", fill="x")

        titulo = ttkbcliner.Label(barra, text="💼 Sistema Financiero Empresarial", bootstyle="inverse-dark", font=("Segoe UI", 12, "bold"))
        titulo.pack(side="left", padx=15)

        usuario_label = ttkbcliner.Label(barra, text=f"👤 {self.usuario}", bootstyle="inverse-dark", font=("Segoe UI", 10))
        usuario_label.pack(side="right", padx=15)

    def mostrar_login(self):
        for widget in self.root.winfo_children():
            widget.destroy()

        self.crear_menu_barra() # Recrear el menú base sin opciones de admin

        marco = ttkbcliner.Frame(self.root, padding=20)
        marco.pack(expand=True)

        ttkbcliner.Label(marco, text="Iniciar sesión", font=("Segoe UI", 14, "bold"), bootstyle="primary").pack(pady=10)
        self.usuario_entry = ttkbcliner.Entry(marco, width=30)
        self.usuario_entry.pack(pady=5)
        self.usuario_entry.insert(0, "Usuario")

        self.contrasena_entry = ttkbcliner.Entry(marco, show="*", width=30)
        self.contrasena_entry.pack(pady=5)
        self.contrasena_entry.insert(0, "")

        ttkbcliner.Button(marco, text="Entrar", bootstyle="success", command=self.login).pack(pady=20)

    def login(self):
        usuario = self.usuario_entry.get()
        contrasena = self.contrasena_entry.get()

        if usuario in usuarios and usuarios[usuario] == contrasena:
            self.usuario, self.tipo = usuario, "usuario"
        elif usuario in administrador and administrador[usuario] == contrasena:
            self.usuario, self.tipo = usuario, "admin"
        else:
            messagebox.showerror("Error de login", "Credenciales incorrectas")
            return

        registrar_log_sesion(self.usuario, self.tipo)
        self.crear_menu_barra()  # actualizar menú si es admin
        self.mostrar_menu_principal()

    def mostrar_menu_principal(self):
        for widget in self.root.winfo_children():
            widget.destroy()

        self.crear_barra_tareas()

        nombre = nombres_usuarios.get(self.usuario, self.usuario)
        ttkbcliner.Label(self.root, text=f"✅ Bienvenida a {self.usuario} ({nombre})", font=("Segoe UI", 12, "bold"), bootstyle="info").pack(pady=10)

        marco = ttkbcliner.Frame(self.root, padding=20)
        marco.pack()

        def crear_boton(texto, comando, estilo="primary"):
            ttkbcliner.Button(marco, text=texto, command=comando, bootstyle=estilo, width=35).pack(pady=5)

        if self.tipo == "admin":
            crear_boton("⚙️ Gestión de Empresas", self.gestion_empresas, "secondary")
            crear_boton("📤 Exportar todos los datos", self.mostrar_vista_previa_exportar, "warning")
            crear_boton("🕒 Ver historial de sesiones", self.ver_logs, "secondary")
        
        elif self.tipo == "usuario":
            crear_boton("🧾 Registrar Gasto", self.mostrar_menu_gastos, "danger")
            crear_boton("💰 Registrar Ingreso", self.mostrar_registrar_ingreso, "success")
            crear_boton("📊 Crear informe por período", self.mostrar_menu_exportacion_usuario, "info")
            crear_boton("📤 Exportar todos mis datos", self.exportar_todo_usuario_actual, "warning")

        ttkbcliner.Button(self.root, text="🔒 Cerrar sesión", command=self.mostrar_login, bootstyle="danger", width=35).pack(pady=20)

    def mostrar_menu_gastos(self):
        ventana = Toplevel(self.root)
        ventana.title("Registrar Gastos Detallados")
        ventana.geometry("700x550")

        notebook = ttk.Notebook(ventana)
        notebook.pack(expand=1, fill='both', padx=10, pady=10)

        gastos_fijos_campos = ["Renta", "Luz", "Internet/Teléfono", "Flota", "Transporte", "Limpieza", "Publicidad", "Combustible"]
        gastos_operacionales_campos = ["TSS", "INFOTEP", "Anticipo"]
        gastos_varios_campos = ["Mobiliario", "Consumible/compra","Materiales","equipos" "Otros"]

        frame_fijos = ttkbcliner.Frame(notebook, padding=10)
        frame_operacionales = ttkbcliner.Frame(notebook, padding=10)
        frame_varios = ttkbcliner.Frame(notebook, padding=10)

        notebook.add(frame_fijos, text="Gastos Fijos")
        notebook.add(frame_operacionales, text="Gastos Operacionales")
        notebook.add(frame_varios, text="Gastos Variados")

        def crear_campos(frame, campos):
            entradas = {}
            for campo in campos:
                fila = ttkbcliner.Frame(frame)
                fila.pack(fill='x', pady=4)

                ttkbcliner.Label(fila, text=campo + " (Monto):", width=20).pack(side='left')
                entrada_monto = ttkbcliner.Entry(fila, width=12)
                entrada_monto.pack(side='left', padx=5)

                ttkbcliner.Label(fila, text="Descripción:", width=12).pack(side='left')
                entrada_desc = ttkbcliner.Entry(fila)
                entrada_desc.pack(side='left', fill='x', expand=True)

                entradas[campo] = (entrada_monto, entrada_desc)
            return entradas

        self.entradas_fijos = crear_campos(frame_fijos, gastos_fijos_campos)
        self.entradas_operacionales = crear_campos(frame_operacionales, gastos_operacionales_campos)
        self.entradas_varios = crear_campos(frame_varios, gastos_varios_campos)

        btn_guardar = ttkbcliner.Button(ventana, text="Guardar Gastos", bootstyle="success", command=self.guardar_gastos_detallados)
        btn_guardar.pack(pady=10)
        

    def guardar_gastos_detallados(self):
        registros = []

        def recolectar(entradas, categoria_general):
            for campo, par in entradas.items():
                if isinstance(par, tuple) and len(par) == 2:
                    entrada_monto, entrada_desc = par
                    valor = entrada_monto.get().strip()
                    descripcion = entrada_desc.get().strip()
                else:
                    entrada_monto = par
                    valor = entrada_monto.get().strip()
                    descripcion = ""
                if valor:
                    try:
                        monto = float(valor)
                    except ValueError:
                        messagebox.showwarning("Valor inválido", f"El valor de '{campo}' no es válido")
                        return False
                    fecha = datetime.now().strftime("%Y-%m-%d")
                    registros.append([fecha, self.usuario, monto, f"{categoria_general} - {campo}", descripcion])
            return True

        if not recolectar(self.entradas_fijos, "Gastos Fijos"): return
        if not recolectar(self.entradas_operacionales, "Gastos Operacionales"): return
        if not recolectar(self.entradas_varios, "Gastos Variados"): return
        if not registros:
            messagebox.showwarning("Sin datos", "No se ingresaron gastos válidos")
            return

        if self.tipo == "admin":
            for fila in registros:
                guardar_csv(GASTOS_FILE, fila)
                guardar_en_mysql("Gastos", fila)
            messagebox.showinfo("Éxito", "Gastos guardados correctamente")
            return

        columnas = ["Fecha", "Usuario", "Monto", "Categoría", "Descripción"]

        def validar(dataset):
            for i, fila in enumerate(dataset, 1):
                if not fila[0]:
                    return f"Fila {i}: la fecha es obligatoria"
                try:
                    float(str(fila[2]).strip())
                except:
                    return f"Fila {i}: el monto debe ser numérico"
                if not fila[3]:
                    return f"Fila {i}: la categoría es obligatoria"
            return True

        def confirmar_guardado(dataset):
            for fila in dataset:
                fila[2] = float(str(fila[2]).strip())
                guardar_csv(GASTOS_FILE, fila)
                guardar_en_mysql("Gastos", fila)
            messagebox.showinfo("Éxito", "Gastos guardados correctamente")

        # Vista previa antes de guardar, usando VistaPrevia del preview_module
        VistaPrevia(
            master=self.root,
            columns=columnas,
            data=registros,
            on_save=confirmar_guardado,
            validation_func=validar,
            title="Vista previa de gastos"
        )


    def mostrar_registrar_ingreso(self):
        ventana = Toplevel(self.root)
        ventana.title("Registrar Ingreso")
        ventana.geometry("400x420")

        ttkbcliner.Label(ventana, text="Monto del ingreso:", font=("Segoe UI", 12)).pack(pady=8)
        monto_entry = ttkbcliner.Entry(ventana)
        monto_entry.pack(pady=5, fill='x', padx=20)

        ttkbcliner.Label(ventana, text="Fuente del ingreso:", font=("Segoe UI", 12)).pack(pady=8)
        fuente_var = StringVar(value="Transferencia")
        opciones_fuente = ["Transferencia", "Tarjeta", "Efectivo"]
        fuente_menu = ttkbcliner.OptionMenu(ventana, fuente_var, opciones_fuente[0], *opciones_fuente)
        fuente_menu.pack(pady=5, fill='x', padx=20)

        ttkbcliner.Label(ventana, text="Descripción:", font=("Segoe UI", 12)).pack(pady=8)
        descripcion_text = Text(ventana, height=4)
        descripcion_text.pack(pady=5, fill='both', padx=20)

        def guardar_ingreso():
            monto_str = monto_entry.get().strip()
            if not monto_str:
                messagebox.showwarning("Entrada requerida", "Ingrese un monto")
                return
            try:
                monto = float(monto_str)
            except ValueError:
                messagebox.showwarning("Valor inválido", "El monto debe ser un número")
                return

            fuente = fuente_var.get()
            descripcion = descripcion_text.get("1.0", "end").strip()
            fecha = datetime.now().strftime("%Y-%m-%d")
            datos = [fecha, self.usuario, monto, fuente, descripcion]

            if self.tipo == "admin":
                guardar_csv(INGRESOS_FILE, datos)
                guardar_en_mysql("Ingresos", datos)
                messagebox.showinfo("Éxito", "Ingreso registrado exitosamente")
                ventana.destroy()
                return

            columnas = ["Fecha", "Usuario", "Monto", "Categoría", "Descripción"]

            def validar(dataset):
                fila = dataset[0]
                if not fila[0]:
                    return "La fecha es obligatoria"
                try:
                    float(str(fila[2]).strip())
                except:
                    return "El monto debe ser numérico"
                if not fila[3]:
                    return "La categoría/fuente es obligatoria"
                return True

            def confirmar_guardado(dataset):
                fila = dataset[0]
                fila[2] = float(str(fila[2]).strip())
                guardar_csv(INGRESOS_FILE, fila)
                guardar_en_mysql("Ingresos", fila)
                messagebox.showinfo("Éxito", "Ingreso registrado exitosamente")
                ventana.destroy()

            VistaPrevia(
                master=self.root,
                columns=columnas,
                data=[datos],
                on_save=confirmar_guardado,
                validation_func=validar,
                title="Vista previa de ingreso"
            )

        ttkbcliner.Button(ventana, text="Guardar Ingreso", bootstyle="success", command=guardar_ingreso).pack(pady=15)



    def mostrar_vista_previa_exportar(self):
        df_gastos = pd.DataFrame(leer_csv(GASTOS_FILE))
        df_ingresos = pd.DataFrame(leer_csv(INGRESOS_FILE))
        if df_gastos.empty and df_ingresos.empty:
            messagebox.showerror("Error", "No hay datos para mostrar")
            return
        # VistaPrevia(self.root, df_gastos, df_ingresos, self.exportar_todo_excel)
        # Implementación simple de vista previa en una ventana Toplevel
        ventana = Toplevel(self.root)
        ventana.title("Vista Previa de Exportación")
        ventana.geometry("800x600")

        text_area = Text(ventana, wrap="word", font=("Courier New", 10))
        text_area.pack(expand=True, fill="both", padx=10, pady=10)

        if not df_gastos.empty:
            text_area.insert("end", "=== GASTOS ===\n")
            text_area.insert("end", df_gastos.to_string(index=False))
            text_area.insert("end", "\n\n")
        if not df_ingresos.empty:
            text_area.insert("end", "=== INGRESOS ===\n")
            text_area.insert("end", df_ingresos.to_string(index=False))
            text_area.insert("end", "\n\n")

        def exportar_y_cerrar():
            self.exportar_todo_excel()
            ventana.destroy()

        ttkbcliner.Button(ventana, text="Exportar a Excel", bootstyle="success", command=exportar_y_cerrar).pack(pady=10)

    def exportar_todo_excel(self):
        exportar_a_excel_consolidado(GASTOS_FILE, INGRESOS_FILE, "datos_consolidados.xlsx", usuario_actual=self.usuario)

    def ver_logs(self):
        if not os.path.exists(LOG_SESIONES):
            messagebox.showerror("Error", "No hay registros de sesiones")
            return
        logs = leer_csv(LOG_SESIONES)
        log_text = "\n".join([" | ".join(row) for row in logs])
        
        # Crear una ventana Toplevel para mostrar los logs
        ventana_logs = Toplevel(self.root)
        ventana_logs.title("Historial de Sesiones")
        ventana_logs.geometry("500x400")
        
        text_area = Text(ventana_logs, wrap="word", font=("Courier New", 10))
        text_area.pack(expand=True, fill="both", padx=10, pady=10)
        text_area.insert("1.0", log_text)
        text_area.config(state="disabled") # Hacer el texto de solo lectura
        
        ttk.Scrollbar(text_area, command=text_area.yview).pack(side="right", fill="y")
        text_area['yscrollcommand'] = ttk.Scrollbar(text_area).set

    def mostrar_menu_exportacion_usuario(self):
        ventana = Toplevel(self.root)
        ventana.title("Crear Informe por Período")
        ventana.geometry("400x220")

        ttkbcliner.Label(ventana, text="Seleccione el período:", font=("Segoe UI", 12)).pack(pady=15)

        periodo_var = StringVar(value="semanal")
        opciones = ["semanal", "mensual"]
        periodo_menu = ttkbcliner.OptionMenu(ventana, periodo_var, opciones[0], *opciones)
        periodo_menu.pack(pady=5)

        def exportar():
            periodo = periodo_var.get()
            self.exportar_reporte_usuario(self.usuario, periodo)
            ventana.destroy()

        ttkbcliner.Button(ventana, text="Exportar Informe", bootstyle="success", command=exportar).pack(pady=20)

    def exportar_reporte_usuario(self, usuario, periodo="semanal"):
        def filtrar_datos_por_periodo(df, periodo):
            if df.empty:
                return df
            
            # Asignar nombres de columna temporales para la manipulación
            temp_cols = [f'col_{i}' for i in range(len(df.columns))]
            df.columns = temp_cols
            
            df[temp_cols[0]] = pd.to_datetime(df[temp_cols[0]], errors='coerce')
            hoy = datetime.now()
            if periodo == "semanal":
                fecha_inicio = hoy - pd.Timedelta(days=hoy.weekday())
            elif periodo == "mensual":
                fecha_inicio = hoy.replace(day=1)
            else:
                return df
            return df[df[temp_cols[0]] >= fecha_inicio]

        df_gastos = pd.DataFrame(leer_csv(GASTOS_FILE))
        df_ingresos = pd.DataFrame(leer_csv(INGRESOS_FILE))

        if not df_gastos.empty:
            df_gastos = df_gastos[df_gastos[1] == usuario]
            df_gastos = filtrar_datos_por_periodo(df_gastos, periodo)

        if not df_ingresos.empty:
            df_ingresos = df_ingresos[df_ingresos[1] == usuario]
            df_ingresos = filtrar_datos_por_periodo(df_ingresos, periodo)

        if df_gastos.empty and df_ingresos.empty:
            messagebox.showinfo("Sin datos", "No hay datos disponibles para exportar en este período.")
            return

        archivo_excel = f"{usuario}_{periodo}_reporte.xlsx"
        exportar_a_excel_consolidado(df_gastos, df_ingresos, archivo_excel, usuario_actual=usuario)

    def exportar_todo_usuario_actual(self):
        df_gastos = pd.DataFrame(leer_csv(GASTOS_FILE))
        df_ingresos = pd.DataFrame(leer_csv(INGRESOS_FILE))

        if not df_gastos.empty:
            df_gastos = df_gastos[df_gastos[1] == self.usuario]
        if not df_ingresos.empty:
            df_ingresos = df_ingresos[df_ingresos[1] == self.usuario]

        if df_gastos.empty and df_ingresos.empty:
            messagebox.showinfo("Sin datos", "No hay datos disponibles para exportar.")
            return

        archivo_excel = f"{self.usuario}_exportacion_completa.xlsx"
        exportar_a_excel_consolidado(df_gastos, df_ingresos, archivo_excel, usuario_actual=self.usuario)
    
    

# --- GESTIÓN DE EMPRESAS (CORREGIDO) ---
    # (Eliminado el uso incorrecto de '...')

    def gestion_empresas(self):
        ventana = Toplevel(self.root)
        ventana.title("Gestión de Empresas")
        ventana.geometry("400x300")

        ttkbcliner.Label(ventana, text="🏢 Selección de Empresa",
                         font=("Segoe UI", 14, "bold")).pack(pady=10)

        for empresa in CLAVES_EMPRESAS.keys():
            ttkbcliner.Button(
                ventana,
                text=empresa,
                command=lambda e=empresa, v=ventana: self.validar_clave_empresa(e, v)
            ).pack(pady=5, padx=20, fill='x')

    def validar_clave_empresa(self, empresa, ventana_padre=None):
        clave = simpledialog.askstring("Clave de Empresa", f"Ingrese la clave para '{empresa}':", show="*")
        if clave == CLAVES_EMPRESAS.get(empresa):
            if ventana_padre is not None:
                ventana_padre.destroy()
            self.mostrar_consultas_empresa(empresa)
        else:
            messagebox.showerror("Error", f"Clave incorrecta para la empresa '{empresa}'")

    def mostrar_consultas_empresa(self, empresa):
        ventana = Toplevel(self.root)
        ventana.title(f"Consultas - {empresa}")
        ventana.geometry("1200x850")

        # --- Filtros ---
        marco_filtros = ttkbcliner.Frame(ventana, padding=10)
        marco_filtros.pack(fill="x")
        
        # año

        ttkbcliner.Label(marco_filtros, text="Año:").pack(side="left", padx=5)
        anios = [str(a) for a in range(2020, 2031)]
        combo_anio = ttkbcliner.Combobox(marco_filtros, values=anios, width=6, state="readonly")
        combo_anio.set("")  # vacío por defecto
        combo_anio.pack(side="left", padx=5)
        entry_anio = combo_anio

        #mes

        ttkbcliner.Label(marco_filtros, text="Mes:").pack(side="left", padx=5)
        meses = [str(m) for m in range(1, 13)]
        combo_mes = ttkbcliner.Combobox(marco_filtros, values=meses, width=4, state="readonly")
        combo_mes.set("")
        combo_mes.pack(side="left", padx=5)
        entry_mes = combo_mes

    

        # Día
        ttkbcliner.Label(marco_filtros, text="Día:").pack(side="left", padx=5)
        dias = [str(d) for d in range(1, 32)]
        combo_dia = ttkbcliner.Combobox(marco_filtros, values=dias, width=4, state="readonly")
        combo_dia.set("")
        combo_dia.pack(side="left", padx=5)
        entry_dia = combo_dia

        # --- Tablas ---
        marco_tablas = ttkbcliner.Frame(ventana, padding=10)
        marco_tablas.pack(fill="both", expand=True)

        ttkbcliner.Label(marco_tablas, text="📉 Gastos").pack()
        tabla_gastos = ttk.Treeview(marco_tablas, columns=("Fecha","Usuario","Monto","Categoría","Descripción"), show="headings")
        for col in ("Fecha","Usuario","Monto","Categoría","Descripción"):
            tabla_gastos.heading(col, text=col)
        tabla_gastos.pack(fill="x", pady=5)

        ttkbcliner.Label(marco_tablas, text="📈 Ingresos").pack()
        tabla_ingresos = ttk.Treeview(marco_tablas, columns=("Fecha","Usuario","Monto","Fuente","Descripción"), show="headings")
        for col in ("Fecha","Usuario","Monto","Fuente","Descripción"):
            tabla_ingresos.heading(col, text=col)
        tabla_ingresos.pack(fill="x", pady=5)

        ttkbcliner.Label(marco_tablas, text="🔥 Gasto más recurrente por mes").pack()
        tabla_recurrentes = ttk.Treeview(marco_tablas, columns=("Mes","Categoría","Repeticiones"), show="headings")
        for col in ("Mes","Categoría","Repeticiones"):
            tabla_recurrentes.heading(col, text=col)
        tabla_recurrentes.pack(fill="x", pady=5)

        # Botón de consulta

        ttkbcliner.Button(
            ventana,
            text="Consultar",
            bootstyle="success",
            command=lambda: self.ejecutar_consulta(
                empresa, combo_anio, combo_mes, combo_dia,
                tabla_gastos, tabla_ingresos, tabla_recurrentes
            )
        ).pack(side="left", padx=10)

# --- Función corregida ---


    def ejecutar_consulta(self, empresa, combo_anio, combo_mes, combo_dia,
                         tabla_gastos, tabla_ingresos, tabla_recurrentes):
        # Limpiar tablas
        for tabla in (tabla_gastos, tabla_ingresos, tabla_recurrentes):
            for row in tabla.get_children():
                tabla.delete(row)

        # Parámetros de filtros
        params = {
            "empresa": empresa,
            "anio": combo_anio.get().strip(),
            "mes": combo_mes.get().strip(),
            "dia": combo_dia.get().strip()
        }

       # Endpoint correcto de la API
        url = "https://api-powerbi-xoh2.onrender.com/api/consultas"  # 👈 cambia por tu URL real
        try:
            response = requests.get(url, params=params, timeout=50, auth=("powerbi", "secure123"))
            response.raise_for_status()

            data = response.json()
            
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo obtener datos de la API: {e}")
            return

        gastos = data.get("gastos", [])
        ingresos = data.get("ingresos", [])
        recurrente = data.get("recurrente", [])

        # Insertar en tablas
        for g in gastos:
            tabla_gastos.insert("", "end", values=(g["fecha"], g["usuario"], g["monto"], g["categoria"], g["descripcion"]))
        for i in ingresos:
            tabla_ingresos.insert("", "end", values=(i["fecha"], i["usuario"], i["monto"], i["fuente"], i["descripcion"]))
        for r in recurrente:
            tabla_recurrentes.insert("", "end", values=(params["mes"], r["categoria"], r["repeticiones"]))




        # --- separador ---

    
# --- Utilidades de archivo ---
def guardar_csv(archivo, datos):
    with open(archivo, mode='a', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(datos)

# (aquí sigue todo tu código igual hasta las funciones que guardan datos)

def guardar_gastos_detallados(self):
        registros = []

        def recolectar(entradas, categoria_general):
            for campo, par in entradas.items():
                if isinstance(par, tuple) and len(par) == 2:
                    entrada_monto, entrada_desc = par
                    valor = entrada_monto.get().strip()
                    descripcion = entrada_desc.get().strip()
                else:
                    entrada_monto = par
                    valor = entrada_monto.get().strip()
                    descripcion = ""
                if valor:
                    try:
                        monto = float(valor)
                    except ValueError:
                        messagebox.showwarning("Valor inválido", f"El valor de '{campo}' no es válido")
                        return False
                    fecha = datetime.now().strftime("%Y-%m-%d")
                    registros.append([fecha, self.usuario, monto, f"{categoria_general} - {campo}", descripcion])
            return True

        if not recolectar(self.entradas_fijos, "Gastos Fijos"): return
        if not recolectar(self.entradas_operacionales, "Gastos Operacionales"): return
        if not recolectar(self.entradas_varios, "Gastos Variados"): return
        if not registros:
            messagebox.showwarning("Sin datos", "No se ingresaron gastos válidos")
            return

        if self.tipo == "admin":
            for fila in registros:
                guardar_csv(GASTOS_FILE, fila)
                guardar_en_mysql("Gastos", fila)
            messagebox.showinfo("Éxito", "Gastos guardados correctamente")
            return

        columnas = ["Fecha", "Usuario", "Monto", "Categoría", "Descripción"]

        def validar(dataset):
            for i, fila in enumerate(dataset, 1):
                if not fila[0]:
                    return f"Fila {i}: la fecha es obligatoria"
                try:
                    float(str(fila[2]).strip())
                except:
                    return f"Fila {i}: el monto debe ser numérico"
                if not fila[3]:
                    return f"Fila {i}: la categoría es obligatoria"
            return True

        def confirmar_guardado(dataset):
            for fila in dataset:
                fila[2] = float(str(fila[2]).strip())
                guardar_csv(GASTOS_FILE, fila)
                guardar_en_mysql("Gastos", fila)
            messagebox.showinfo("Éxito", "Gastos guardados correctamente")

        VistaPrevia(
            master=self.root,
            columns=columnas,
            data=registros,
            on_save=confirmar_guardado,
            validation_func=validar,
            title="Vista previa de gastos"
        )

def mostrar_registrar_ingreso(self):
        ventana = Toplevel(self.root)
        ventana.title("Registrar Ingreso")
        ventana.geometry("420x420")

        ttkbcliner.Label(ventana, text="Monto del ingreso:", font=("Segoe UI", 12)).pack(pady=8)
        monto_entry = ttkbcliner.Entry(ventana)
        monto_entry.pack(pady=5, fill='x', padx=20)

        ttkbcliner.Label(ventana, text="Fuente del ingreso:", font=("Segoe UI", 12)).pack(pady=8)
        fuente_var = StringVar(value="Transferencia")
        opciones_fuente = ["Transferencia", "Tarjeta", "Efectivo"]
        fuente_menu = ttkbcliner.OptionMenu(ventana, fuente_var, opciones_fuente[0], *opciones_fuente)
        fuente_menu.pack(pady=5, fill='x', padx=20)

        ttkbcliner.Label(ventana, text="Descripción:", font=("Segoe UI", 12)).pack(pady=8)
        descripcion_text = Text(ventana, height=4)
        descripcion_text.pack(pady=5, fill='both', padx=20)

        def guardar_ingreso():
            monto_str = monto_entry.get().strip()
            if not monto_str:
                messagebox.showwarning("Entrada requerida", "Ingrese un monto")
                return
            try:
                monto = float(monto_str)
            except ValueError:
                messagebox.showwarning("Valor inválido", "El monto debe ser un número")
                return

            fuente = fuente_var.get()
            descripcion = descripcion_text.get("1.0", "end").strip()
            fecha = datetime.now().strftime("%Y-%m-%d")
            datos = [fecha, self.usuario, monto, fuente, descripcion]

            if self.tipo == "admin":
                guardar_csv(INGRESOS_FILE, datos)
                guardar_en_mysql("Ingresos", datos)
                messagebox.showinfo("Éxito", "Ingreso registrado exitosamente")
                ventana.destroy()
                return

            columnas = ["Fecha", "Usuario", "Monto", "Categoría", "Descripción"]

            def validar(dataset):
                fila = dataset[0]
                if not fila[0]:
                    return "La fecha es obligatoria"
                try:
                    float(str(fila[2]).strip())
                except:
                    return "El monto debe ser numérico"
                if not fila[3]:
                    return "La categoría/fuente es obligatoria"
                return True

            def confirmar_guardado(dataset):
                fila = dataset[0]
                fila[2] = float(str(fila[2]).strip())
                guardar_csv(INGRESOS_FILE, fila)
                guardar_en_mysql("Ingresos", fila)
                messagebox.showinfo("Éxito", "Ingreso registrado exitosamente")
                ventana.destroy()

            VistaPrevia(
                master=self.root,
                columns=columnas,
                data=[datos],
                on_save=confirmar_guardado,
                validation_func=validar,
                title="Vista previa de ingreso"
            )

        ttkbcliner.Button(ventana, text="Guardar Ingreso", bootstyle="success", command=guardar_ingreso).pack(pady=15)

# --- EJECUCIÓN ---
if __name__ == "__main__":
    app = ttkbcliner.Window(themename="morph")
    SistemaFinancieroApp(app)
    app.mainloop()

